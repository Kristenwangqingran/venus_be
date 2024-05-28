# 2021-03-31 minhao.zhang created
# 2021-06-28 minhao.zhang add xnode2xmind
import os
import json
from pathlib import Path
from collections import OrderedDict

import xmind
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, named_styles
from pprint import pprint

"""总体解析思路 parse() 下的逻辑
excel每一行的（有效）数据，都会是xmind上的一条完整的链路，但是同一个单元格如果是一样的话，是可以进行合并的
同时，平台上已有的功能是把xmind_content(一种字典+列表的混合结构)转换成node数据，因此本模块做的事情是
把excel格式转化成xmind_content格式

首先要把excel转成纯字典格式（这样才能便于快速合并相同单元格，实现逻辑在：_parse_row_to_dict
然后进行合并，实现逻辑在：__merge_topics
最后进行数据转换，实现逻辑在：_parse_dict_to_xmind_content
"""

MERGED_CEll = "MergedCell"
# 用例模板在 https://docs.google.com/spreadsheets/d/1GWNqFF-vMJAIXM1w-e7lXx_XptCZtDXCQ6xkxh_amnY/edit#gid=61919685
# 以下配置均根据模板格式而设定
CASES_SHEET = "Cases"  # 表格名要为Cases（和模板相符）
MIN_ROW, MIN_COL = 6, 2  # 从第6行、第二列开始为数据内容
# 模板中不同item所属的列数
DESC_COL, CATE_START_COL, CATE_END_COL, STEP_COL, EXPECTED_RET_COL, PRIORITY_ROW, RESULT_ROW = 1, 2, 5, 7, 8, 9, 10
VERSION_CELL = "B1"
QA_CELL = "B2"
TOTAL_CELL = "B3"
TITLE_ROW = 4  # 第4行表明每一列是什么标题
# 各title名称
CATEGORY = "Category"
PRECONDITION = "Precondition"
STEPS = "Steps"
RESULT = "Result"
EXPECTED_RESULTS = "Expected Result"
DETAILS = "Details"
PRIORITY = "Priority"
VALID_RESULTS = ["ok", "no"]
# for Xnode2Excel
CATE_ROWS_NUM = 4
KEY = "key"
BLOCKS = "blocks"
DATA = "data"
PARENT_KEY = "parentKey"
SUB_KEYS = "subKeys"
STYLE = "style"
CONTENTSTYLE = "contentStyle"


class Excel2Xmind:
    START_ROW = 4
    CATEGERY_ROWS = []

    def __init__(self, excel_file):
        # assert excel_file.endswith(".xlsx")
        wb = load_workbook(excel_file)
        self.sheet = wb[CASES_SHEET]
        # 严格按照xfile表中的xmind_content来构造, 以便复用已有解析逻辑
        self._last_topics = None  # 用于每次topic生成后，对比上一个topics，看是否需要合并
        self._xmind_content = [
            {"title": "excel2xmind", # 实际上该值不影响最终效果，只是做个标记
             "topic": {
                 "title": Path(excel_file).stem,
                 # "topics": self._last_topics
             },
             "structure": "org.xmind.ui.map.unbalanced"}]

        self._cols_name_index_map, self._cols_index_name_map = self.__parse_cols_map()  # 用于快速定位列名字与对应索引

    def __parse_cols_map(self):
        """
        用于记录category名以及所在的列的索引，便于每一行的判断
        RETURNS:
            cols_name_index_map:
                {"CATEGORY": [1,2],
                 "PRECONDITION": [3]
                 }
            cols_index_name_map:
                {1: "CATEGORY",
                 2: "CATEGORY",
                 3: "PRECONDITION"
                }
        """
        cols_name_index_map = {
            CATEGORY: [],
            PRECONDITION: [],
            STEPS: [],
            EXPECTED_RESULTS: [],
            PRIORITY: [],
            RESULT: []
        }
        cols_index_name_map = {}
        for rows in self.sheet.iter_rows(min_row=TITLE_ROW, max_row=TITLE_ROW):  # 按照case模板来看，第4、5行为title行，表明该列表示什么
            for cell in rows:
                cell_value = self._get_cell_value(cell)
                if cols_name_index_map.get(cell_value) is not None:
                    cols_name_index_map[cell_value].append(cell.column)
                    cols_index_name_map[cell.column] = cell_value

        return cols_name_index_map, cols_index_name_map

    def _get_cell_value(self, cell):
        """
        获取每一个单元格的值
        如果是合并单元格，cell.value的值会为None，此时需要找到真实的值
        """
        if type(cell).__name__ != MERGED_CEll:
            if cell.value is not None:
                try:
                    return str(cell.value).strip()
                except:
                    return cell.value
        else:
            rng = [s for s in self.sheet.merged_cells.ranges if cell.coordinate in s]
            return self.sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng) != 0 else cell.value

    def _parse_cell_value_to_key(self, cell_value: str):
        """
        把单元格的值变成字典中的key
        {"value": None}
        """
        ret = OrderedDict()
        ret[cell_value] = None
        return ret

    def __merge_dict(self, dict1, dict2, tmp_topic=None):
        """WARNING: the value of dict1 will be changed after __merge_topics()
        合并两个topic，当同一层级的key相同时，进行合并
        因为最开始写的时候没有考虑到最外部的key存在不同的情况，所以这种情况需要再传入的dict中再套一层
        {1: dict1}, {1:dict2}这样
        t1 = {"t1.1":
                {"t1.1.1":
                    {"t1.1.2": None
                    },
                "t1.1.3": None
                }
              }

        t2 = {"t1.1":
                {"t1.1.1":
                    {"t1.1.3": None
                    },
                 "t1.1.2": None
                }
              }
        合并后 (合并了t1.1.1下的value)
        t = = {"t1.1":
                {"t1.1.1":
                    {"t1.1.2": None，
                     "t1.1.3": None
                    },
                "t1.1.2": None,
                "t1.1.3": None
                }
              }

        """
        if tmp_topic is None:
            tmp_topic = dict1.copy()

        for topics2_k, topics2_v in dict2.items():
            if dict1.get(topics2_k):
                tmp_topic = dict1[topics2_k]
                self.__merge_dict(dict1[topics2_k], topics2_v, tmp_topic)
            else:
                tmp_topic[topics2_k] = topics2_v
        return tmp_topic

    def _parse_row_to_dict(self, row, index=0):
        """把每一行单元格转化成字典，这么设计是为了方便做merge
        | cate1 | cate2 | cate3.1 | step | expected_result
        ->
        {"cate1":
            {"cate2":
                {"cate3,1":
                    {"step":
                        {"expected_result"
                        }
                    }
                }
            }
        }

        """
        if index == len(row):
            return None

        cell = row[index]
        cell_value = self._get_cell_value(cell)
        if cell_value is not None:
            if index == len(row) - 1:
                return self._parse_cell_value_to_key(cell_value)
            else:
                topics = self._parse_cell_value_to_key(cell_value)
                topics[cell_value] = self._parse_row_to_dict(row=row, index=index + 1)
                return topics
        else:
            return self._parse_row_to_dict(row=row, index=index + 1)

    def _parse_dict_to_xmind_content(self, topics):
        """
        把字典格式的topics，转化成xmind的内容，便于和后台复用已有逻辑
        d = {"sms1": {"sms1.1": None,
             "sms1.2": {"sms1.2.1": None}}}
        ===》
        [{'title': 'sms1',
          'topics': [{'title': 'sms1.1'},
             {'title': 'sms1.2', 'topics': [{'title': 'sms1.2.1'}]}]}]
        """
        ret = []
        for k, v in topics.items():
            if v is None:
                topic = {"title": str(k)}
            else:
                topic = {
                    "title": str(k),
                    "topics": self._parse_dict_to_xmind_content(v)
                }
            ret.append(topic)
        return ret

    def parse(self):
        # 遍历到EXPECTED_RESULTS，因为平台上还不支持打标签之类的功能，没办法标记成功与失败，所以只标记用例
        _tmp = 1  # 用于merge时的临时变量
        for row in self.sheet.iter_rows(min_row=MIN_ROW, min_col=MIN_COL,
                                        max_col=self._cols_name_index_map[EXPECTED_RESULTS][-1]):
            current_topics = self._parse_row_to_dict(row=row)
            if current_topics is not None:
                if self._last_topics is None:
                    self._last_topics = current_topics
                else:
                    self.__merge_dict({_tmp: self._last_topics}, {_tmp: current_topics})
            # break
        topics = self._parse_dict_to_xmind_content(self._last_topics)

        self._xmind_content[0]["topic"]["topics"] = topics
        return self._xmind_content

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass


class StepList(list):
    """自定义step列表
    用于去除空格的情况
    """

    def append(self, value):
        if value == "\n":
            ...
        else:
            super(StepList, self).append(value.strip())


class Xnode2Excel:
    """
    非标准xmind格式文件，而是平台自定义的数据结构，转换成符合当前部门使用的excel模板
    nodes example:
    [{'blocks': [{'data': '4.xmind', 'type': 'CONTENT'}],
      'collapse': False,
      'key': '3ad896c4-ed2e-497c-b82c-d125332959b8',
      'parentKey': None,
      'style': None,
      'subKeys': ['3fd7d4ea-2002-4b04-905c-c0142afebb71',
                  '5b9f0cdb-31d6-4ca1-a612-f5c04baf23b3',
                  '93eec147-1135-4b93-9239-b82c65c0b0b6']},
     {'blocks': [{'data': 'fadsfa', 'type': 'CONTENT'}],
      'collapse': False,
      'key': '3fd7d4ea-2002-4b04-905c-c0142afebb71',
      'parentKey': '3ad896c4-ed2e-497c-b82c-d125332959b8',
      'style': None,
      'subKeys': ['4322c3a4-a031-4c79-9d04-0b336657e98a']},
     {'blocks': [{'data': 'dddd', 'type': 'CONTENT'}],
      'collapse': False,
      'key': '93eec147-1135-4b93-9239-b82c65c0b0b6',
      'parentKey': '3ad896c4-ed2e-497c-b82c-d125332959b8',
      'style': None,
      'subKeys': []},
     {'blocks': [{'data': 'aaaa', 'type': 'CONTENT'}],
      'collapse': False,
      'key': '5b9f0cdb-31d6-4ca1-a612-f5c04baf23b3',
      'parentKey': '3ad896c4-ed2e-497c-b82c-d125332959b8',
      'style': None,
      'subKeys': []},
     {'blocks': [{'data': '发达', 'type': 'CONTENT'}],
      'collapse': False,
      'key': '4322c3a4-a031-4c79-9d04-0b336657e98a',
      'parentKey': '3fd7d4ea-2002-4b04-905c-c0142afebb71',
      'style': None,
      'subKeys': []}]

    """
    TEMPLATE_FILE = Path(__file__).parent.absolute() / "TestCaseTemplate.xlsx"
    # 单元格样式
    # 文字对齐
    _ali = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 边框
    _line = Side(style='thin', color='000000')
    _border = Border(top=_line, bottom=_line, left=_line, right=_line)
    STYLE = named_styles.NamedStyle(name="STYLE", alignment=_ali, border=_border)

    def __init__(self, nodes, output_file, version=None, QA=None):
        self.nodes = nodes
        self.output_file = output_file
        self.version = version
        self.QA = QA
        self._cate_nodes = {}
        self._step_nodes = {}
        self._expected_result_nodes = {}
        self._greatest_common_title = []  # 最大公共title
        self._key_node_map = {node[KEY]: node for node in self.nodes}
        self._xlsx_lines = []  # 用于记录解析node之后，填充excel的每行数据

    def _get_title_from_blocks(self, blocks):
        # "blocks": [{
        #     "type": "CONTENT",
        #     "data": "can go to specific page"
        # },
        #     {
        #     "type": "DESC",
        #     "data": "hi, this is 笔记"
        # },
        #     {
        #     "type": "DESC",
        #     "data": "['priority-1']"
        # }
        # ]
        for the_dict in blocks:
            if the_dict.get('type') == 'CONTENT':
                return the_dict.get('data')
            else:
                continue
        return 'No title'

    def _get_root_node(self):
        for node in self.nodes:
            if not node[PARENT_KEY]:
                return node

    def _parse_node_to_list(self, node=None):
        """该方法核心思路是把带有父子关系的树状node转换成符合excel格式的行数据
        从根节点开始，逐个解析节点title，并且添加到每一行的数据中
        """
        if node is None:
            node = self._get_root_node()
        title = self._get_title_from_blocks(node[BLOCKS])
        sub_keys_num = len(node[SUB_KEYS])
        self._greatest_common_title.append(title)

        if sub_keys_num > 0:
            # 存在子节点的情况下
            # 记录最大公共列
            #pprint(title)
            for sub_key in node[SUB_KEYS]:
                self._parse_node_to_list(node=self._key_node_map[sub_key])

            self._greatest_common_title.pop()
        else:
            line = self._greatest_common_title.copy()
            self._greatest_common_title.pop()
            # 记录result
            try:
                style = json.loads(node.get(STYLE))
                result = style.get(CONTENTSTYLE, {}).get("result", "")
            except Exception:
                result = ""
            line.append(result)
            self._xlsx_lines.append(line)

        return self._xlsx_lines

    def _to_excel(self, lines_content):
        """the structure of lines_content:
        [['4.xmind', 'fadsfa', '发达'],
         ['4.xmind', 'aaaa'],
         ['4.xmind', 'dddd']]
        """
        # pprint(lines_content)
        wb = load_workbook(self.TEMPLATE_FILE)
        sheet = wb[CASES_SHEET]
        # basic info
        sheet[VERSION_CELL] = self.version
        sheet[QA_CELL] = self.QA
        sheet[TOTAL_CELL] = len(self._xlsx_lines)

        all_cases_length = len(lines_content)
        # case info
        for row_index, line in enumerate(lines_content):
            result = line.pop()  # 先取出result，避免影响原有的逻辑
            # index+MIN_ROW 表示当前填充数据的行数
            length = len(line)
            step_content = StepList()
            for col_index, title in enumerate(line):
                # index表示当前填充数据的列数
                # ['4.xmind', 'fadsfa', '发达']
                if CATE_ROWS_NUM < col_index < length-1:
                    # index大于4的节点，往step里面放(因为Category最多为4列)
                    step_content.append(title.rstrip())
                    continue
                if col_index == 0:
                    # 根节点
                    # sheet.cell(row=MIN_ROW+row_index, column=CATE_START_COL, value=title)
                    continue
                elif col_index == length-1:
                    # 最尾节点必为expected output
                    sheet.cell(row=MIN_ROW+row_index, column=EXPECTED_RET_COL, value=title)
                    # 保存之前的step_content
                    sheet.cell(row=MIN_ROW+row_index, column=STEP_COL, value="\n".join(step_content))
                    # 设置每个case的Priority（默认为1）
                    sheet.cell(row=MIN_ROW+row_index, column=PRIORITY_ROW, value=1)
                elif col_index == length-2:
                    # 倒数第二节点必为step
                    step_content.append(title.rstrip())
                else:
                    # 其他情况，优先填满Category，多的拼接到step
                    if col_index < CATE_ROWS_NUM+1:
                        current_row = MIN_ROW + row_index
                        current_column = CATE_START_COL + col_index - 1  # 因为第一个节点不放在cate中，所以这里需要往前挪一列，即 -1
                        cell = sheet.cell(row=current_row, column=current_column, value=title)
                        # cell.style = self.STYLE
                    else:
                        step_content.append(title.rstrip())
                # 最后填充结果
                sheet.cell(row=MIN_ROW + row_index, column=RESULT_ROW, value=result)

        # merge sheet
        for column in sheet.iter_cols(min_row=MIN_ROW, max_row=MIN_ROW + all_cases_length - 1, min_col=CATE_START_COL, max_col=CATE_END_COL):
            start_row = MIN_ROW
            end_row = MIN_ROW
            for cell in column:
                current_value = cell.value

                last_value = sheet.cell(row=cell.row-1, column=cell.column).value

                if current_value == last_value:
                    end_row = cell.row
                else:
                    # 合并
                    sheet.merge_cells(start_row=start_row, start_column=cell.column,
                                      end_row=end_row, end_column=cell.column)
                    start_row = cell.row
                    end_row = cell.row
            sheet.merge_cells(start_row=start_row, start_column=cell.column,
                              end_row=end_row, end_column=cell.column)

        # set style
        for index, rows in enumerate(sheet.iter_rows(min_row=MIN_ROW, max_row=MIN_ROW+len(self._xlsx_lines)-1)):
            length = max(len(str(cell.value)) for cell in rows if cell.value is not None)
            sheet.row_dimensions[index + MIN_ROW].height = length + 10
            for cell in rows:
                cell.style = self.STYLE

        # for column_cells in sheet.columns:
        #     length = max(len(cell.value) for cell in column_cells if cell.value is not None)
        #     sheet.column_dimensions[column_cells[0].column].width = length

        wb.save(self.output_file)

    def parse(self):
        xlsx_lines = self._parse_node_to_list()
        # 执行完上述方法后，self._xlsx_lines 已经填充完数据
        # exp:
        # [['4.xmind', 'fadsfa', '发达'],
        #  ['4.xmind', 'aaaa'],
        #  ['4.xmind', 'dddd']]
        self._to_excel(lines_content=xlsx_lines)


class Xnode2Xmind8:
    """把当前平台存入的数据结构转换成xmind文件
    这里有很多通用逻辑，与Node2Excel相同，有时间的话可以整合这里的代码
    """
    REG_ICON = {"Yes": "task-done",
                "No": "task-start"}
    CASE_TYPE_ICON = {"APP": "flag-red",
                      "API": "flag-yellow",
                      "UI": "flag-blue",
                      "Others": "flag-dark-gray"}
    AUTO_ICON = {"Yes": "star-green",
                 "No": "star-red"}
    PRIORITY_ICON = {
        "Priority-1": "priority-1"
    }

    def __init__(self, nodes: list, output_file: str):
        self.nodes = nodes
        self.output_file = output_file if output_file.endswith(".xmind") else f"{output_file}.xmind"
        if os.path.exists(self.output_file):
            os.remove(self.output_file)
        self.webhook = None

        self._key_node_map = {node[KEY]: node for node in self.nodes}

    def _get_title_from_blocks(self, blocks):
        # "blocks": [{
        #     "type": "CONTENT",
        #     "data": "can go to specific page"
        # },
        #     {
        #     "type": "DESC",
        #     "data": "hi, this is 笔记"
        # },
        #     {
        #     "type": "DESC",
        #     "data": "['priority-1']"
        # }
        # ]
        for block in blocks:
            if block.get('type') == 'CONTENT':
                return block.get('data')
            else:
                continue
        return 'No title'

    def _get_root_node(self):
        for node in self.nodes:
            if not node[PARENT_KEY]:
                return node

    def _set_topic_icon(self, node, topic):
        style = node[STYLE]
        if style:
            # add icon
            style = json.loads(style).get('contentStyle', {})
            topic.addMarker(self.REG_ICON.get(style.get("reg")))
            topic.addMarker(self.CASE_TYPE_ICON.get(style.get("type")))
            topic.addMarker(self.AUTO_ICON.get(style.get("auto")))
        # default priority is 1
        topic.addMarker(self.PRIORITY_ICON['Priority-1'])

    def _parse_node_to_xmind(self, node, topic):
        """main logic of parsing node to xmind8
        """
        topic.setTitle(self._get_title_from_blocks(node[BLOCKS]))

        if not node[SUB_KEYS]:
            # last step
            self._set_topic_icon(node, topic)
            return topic
        else:
            for sub_key in node[SUB_KEYS]:
                sub_topic = topic.addSubTopic()
                sub_node = self._key_node_map[sub_key]
                self._parse_node_to_xmind(node=sub_node, topic=sub_topic)

    def _parse(self):
        sheet = self.webhook.getPrimarySheet()
        root_node = self._get_root_node()
        sheet.setTitle(f"{self.output_file}")

        root_topic = sheet.getRootTopic()

        self._parse_node_to_xmind(node=root_node, topic=root_topic)

    def parse(self):
        self.webhook = xmind.load(f"{self.output_file}")
        self._parse()
        xmind.save(self.webhook)


def excel_to_dict(excel_file):
    excel2xmind = Excel2Xmind(excel_file)
    return excel2xmind.parse()


if __name__ == "__main__":
    from pprint import pprint
    # import argparse
    # parser = argparse.ArgumentParser(description='检查excel文件是否能被平台正确解析')
    # parser.add_argument("filename")
    # args = parser.parse_args()
    #
    # t = excel_to_dict(args.filename)
    t = Excel2Xmind("/Users/minhao.zhang/Downloads/中心主题.xmind")
    pprint(t.parse())
    # t = Node2Xmind(nodes=[], title="mmmmmhhhhh")
    # t.parse()
